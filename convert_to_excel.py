# -*- coding: utf-8 -*-
"""
将BTC信号CSV文件转换为Excel格式（带格式化）
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_formatted_excel(csv_file, excel_file):
    """
    创建带格式的Excel文件
    """
    print(f"读取CSV文件: {csv_file}")
    df = pd.read_csv(csv_file)

    # 首先保存为Excel
    print(f"保存为Excel文件: {excel_file}")
    df.to_excel(excel_file, index=False, engine='openpyxl')

    # 加载Excel文件进行格式化
    print("应用格式化...")
    wb = load_workbook(excel_file)
    ws = wb.active

    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    trade_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # 绿色 - 开单
    no_trade_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 红色 - 过滤
    neutral_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色 - 震荡

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 设置列宽
    column_widths = {
        'A': 20,  # 时间
        'B': 12,  # 开盘价
        'C': 12,  # 收盘价
        'D': 12,  # 最高价
        'E': 12,  # 最低价
        'F': 12,  # 交易量
        'G': 12,  # 量能比率
        'H': 12,  # EMA偏离%
        'I': 10,  # 张力
        'J': 10,  # 加速度
        'K': 20,  # 信号类型
        'L': 10,  # 置信度
        'M': 30,  # 信号描述
        'N': 12,  # 交易方向
        'O': 15,  # 通过V705过滤器
        'P': 30,  # 过滤原因
        'Q': 10,  # 是否开单
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 格式化表头
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # 格式化数据行
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # 获取"是否开单"列的值（Q列）
        is_trade = row[16].value  # Q列（第17列）

        # 根据是否开单设置行颜色
        if is_trade == '是':
            fill = trade_fill
            font = Font(color="000000")
        elif is_trade == '否':
            fill = no_trade_fill
            font = Font(color="FFFFFF")
        else:
            fill = neutral_fill
            font = Font(color="000000")

        for cell in row:
            # 设置单元格边框和对齐
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            # 只对非价格数据设置背景色
            if cell.column not in [2, 3, 4, 5]:  # B-E列（价格列）不设置背景
                cell.fill = fill
                cell.font = font

            # 特殊格式化
            if cell.column == 1:  # 时间列
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif cell.column in [2, 3, 4, 5]:  # 价格列
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif cell.column == 6:  # 交易量
                cell.number_format = '#,##0.000'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif cell.column in [7, 8, 9, 10, 12]:  # 数字列
                cell.number_format = '0.000'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif cell.column == 11 or cell.column == 14:  # 信号类型和方向
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif cell.column == 17:  # 是否开单
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 添加筛选器
    ws.auto_filter.ref = ws.dimensions

    # 保存
    wb.save(excel_file)
    print(f"Excel文件已保存: {excel_file}")

    # 统计信息
    print("\n" + "="*70)
    print("数据统计:")
    print(f"  总行数: {len(df)}")
    print(f"  开单信号: {len(df[df['是否开单'] == '是'])}")
    print(f"  过滤信号: {len(df[df['是否开单'] == '否'])}")
    print("\n信号类型分布:")
    for signal_type, count in df['信号类型'].value_counts().items():
        print(f"  {signal_type}: {count}个")
    print("="*70)

    return excel_file


def create_summary_sheet(excel_file, csv_file):
    """
    添加统计摘要sheet
    """
    print("创建统计摘要...")
    df = pd.read_csv(csv_file)

    wb = load_workbook(excel_file)

    # 创建摘要sheet
    if "摘要" in wb.sheetnames:
        wb.remove(wb["摘要"])
    ws_summary = wb.create_sheet("摘要", 0)

    # 统计数据
    total_signals = len(df)
    trade_signals = df[df['是否开单'] == '是']
    filtered_signals = df[df['是否开单'] == '否']

    # 按信号类型统计
    signal_counts = df.groupby(['信号类型', '是否开单']).size().unstack(fill_value=0)

    # 按交易方向统计
    direction_counts = df[df['是否开单'] == '是的'].groupby('交易方向').size() if len(trade_signals) > 0 else pd.Series()

    # 价格统计
    price_stats = df['收盘价'].describe()

    # 写入摘要
    ws_summary['A1'] = 'BTC 4H信号数据统计摘要'
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:D1')

    row = 3
    ws_summary[f'A{row}'] = '数据时间范围'
    ws_summary[f'B{row}'] = f'{df['时间'].min()} 至 {df['时间'].max()}'
    row += 2

    ws_summary[f'A{row}'] = '总体统计'
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws_summary[f'A{row}'] = '总信号数'
    ws_summary[f'B{row}'] = total_signals
    row += 1

    ws_summary[f'A{row}'] = '开单信号'
    ws_summary[f'B{row}'] = len(trade_signals)
    ws_summary[f'B{row}'].font = Font(color="00B050", bold=True)
    row += 1

    ws_summary[f'A{row}'] = '过滤信号'
    ws_summary[f'B{row}'] = len(filtered_signals)
    ws_summary[f'B{row}'].font = Font(color="FF0000", bold=True)
    row += 1

    ws_summary[f'A{row}'] = '开单率'
    ws_summary[f'B{row}'] = f'{len(trade_signals)/total_signals*100:.1f}%'
    row += 3

    # 信号类型分布
    ws_summary[f'A{row}'] = '信号类型分布'
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws_summary[f'A{row}'] = '信号类型'
    ws_summary[f'B{row}'] = '开单'
    ws_summary[f'C{row}'] = '过滤'
    ws_summary[f'D{row}'] = '总计'
    for col in ['A', 'B', 'C', 'D']:
        ws_summary[f'{col}{row}'].font = Font(bold=True)
    row += 1

    for signal_type in signal_counts.index:
        ws_summary[f'A{row}'] = signal_type
        ws_summary[f'B{row}'] = signal_counts.loc[signal_type, '是'] if '是' in signal_counts.columns else 0
        ws_summary[f'C{row}'] = signal_counts.loc[signal_type, '否'] if '否' in signal_counts.columns else 0
        ws_summary[f'D{row}'] = signal_counts.loc[signal_type].sum()
        row += 1
    row += 3

    # 价格统计
    ws_summary[f'A{row}'] = '价格统计（收盘价）'
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws_summary[f'A{row}'] = '最高价'
    ws_summary[f'B{row}'] = f'${price_stats["max"]:,.2f}'
    row += 1

    ws_summary[f'A{row}'] = '最低价'
    ws_summary[f'B{row}'] = f'${price_stats["min"]:,.2f}'
    row += 1

    ws_summary[f'A{row}'] = '平均价'
    ws_summary[f'B{row}'] = f'${price_stats["mean"]:,.2f}'
    row += 1

    ws_summary[f'A{row}'] = '中位数'
    ws_summary[f'B{row}'] = f'${price_stats["50%"]:,.2f}'
    row += 3

    # 说明
    ws_summary[f'A{row}'] = '说明'
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    explanations = [
        "• 绿色行 = 开单信号（通过V7.0.5过滤器）",
        "• 红色行 = 过滤信号（未通过V7.0.5过滤器）",
        "• BEARISH_SINGULARITY = 奇点看空 → 做多",
        "• HIGH_OSCILLATION = 高位震荡 → 做空",
        "• OSCILLATION = 平衡震荡 → 不交易",
        "",
        "• 止盈: +5%",
        "• 止损: -2.5%",
        "• 最大持仓: 42个4H周期（7天）"
    ]

    for explanation in explanations:
        ws_summary[f'A{row}'] = explanation
        row += 1

    # 设置列宽
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 15

    wb.save(excel_file)
    print("统计摘要已添加")


def main():
    """主函数"""
    csv_file = 'btc_4h_signals_complete_20251201_20260119.csv'
    excel_file = 'btc_4h_signals_complete_20251201_20260119.xlsx'

    print("="*70)
    print("BTC 4H信号数据 - Excel生成工具")
    print("="*70)

    # 创建格式化的Excel
    create_formatted_excel(csv_file, excel_file)

    # 添加摘要sheet
    create_summary_sheet(excel_file, csv_file)

    print("\n" + "="*70)
    print(f"Excel文件生成完成!")
    print(f"文件路径: {excel_file}")
    print("="*70)


if __name__ == "__main__":
    main()
